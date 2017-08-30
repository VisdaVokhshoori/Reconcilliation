#!/usr/bin/env python
import sys
import openpyxl

# print (sys.argv[0]);
# print (sys.argv[1]);

# Variable naming convention: CamelCase
# 1-space on each side of mathematical symbols

# Take the xlsx file name and location
Path = sys.argv[1];
WorkBook = openpyxl.load_workbook(Path);

# Load the two sheets
LedgerSheet = WorkBook.get_sheet_by_name('General Ledger');
BankSheet = WorkBook.get_sheet_by_name('Bank Statement');

NrowsLedgerSheet = LedgerSheet.max_row;
NrowsBankSheet = BankSheet.max_row;

# Assumed the two sheets format is like Simplified Data.xlsx
# x iterates over General Ledger row=x,column=2
# y iterates over Bank Statement row=y,column=3
for x in range(2,NrowsLedgerSheet):
  AmountLedgerSheet = LedgerSheet.cell(row=x,column=2).value;
  for y in range(3,NrowsBankSheet):
    AmountBankSheet = BankSheet.cell(row=y,column=3).value;
    if AmountLedgerSheet == AmountBankSheet:
      print ("Perfect match Bank Statement ",y, end=" "); 
      print (" and General Ledger ",x);
  